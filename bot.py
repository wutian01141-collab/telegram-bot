import os
import re
import html
import sqlite3
import logging
from io import BytesIO
from zoneinfo import ZoneInfo
from datetime import datetime, time, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from telegram import Update, KeyboardButton, ReplyKeyboardMarkup, InputFile
from telegram.constants import ParseMode, ChatType
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ContextTypes,
    Defaults,
    filters,
)

# =========================
# 基础配置
# =========================
BOT_TOKEN = os.environ.get("BOT_TOKEN", "").strip()
if not BOT_TOKEN:
    raise ValueError("缺少环境变量 BOT_TOKEN")

DB_FILE = "enterprise_checkin.db"
LOCAL_TZ = ZoneInfo("Asia/Bangkok")

# 可打卡时间：20:00 - 次日12:00
CHECK_START = time(20, 0)
CHECK_END = time(12, 0)

# 正式上下班时间：21:00 - 次日10:00
WORK_START = time(21, 0)
WORK_END = time(10, 0)

# 每天中午12:00切换统计周期，并自动导出上一周期
RESET_TIME = time(12, 0)

# 晚上21:10检查未打上班卡成员
ABSENCE_CHECK_TIME = time(21, 10)

# 导出只允许管理员
EXPORT_ADMIN_ONLY = True

# 超时前多久提醒本人（秒）
WARNING_BEFORE_SECONDS = 60

# 离岗时限（分钟）
DEFAULT_TIMEOUTS = {
    "吃饭": 15,
    "上厕所": 10,
    "抽烟": 10,
}

BREAK_ACTIONS = ["吃饭", "上厕所", "抽烟"]

logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("th-enterprise-checkin-bot")


# =========================
# 中英识别
# =========================
ACTION_KEYWORDS = {
    "上班": [
        "上班", "到岗", "到了", "开工", "打卡上班",
        "on", "on duty", "clock in", "start work", "check in",
    ],
    "下班": [
        "下班", "收工", "走了", "打卡下班",
        "off", "off duty", "clock out", "off work", "check out",
    ],
    "吃饭": [
        "吃饭", "去吃饭", "干饭", "吃个饭", "吃东西",
        "eat", "eating", "meal", "lunch", "dinner", "go eat",
    ],
    "上厕所": [
        "上厕所", "厕所", "洗手间", "卫生间", "方便一下",
        "toilet", "bathroom", "restroom", "wc", "washroom",
    ],
    "抽烟": [
        "抽烟", "去抽烟", "来一根", "吸烟",
        "smoke", "smoking", "cigarette", "go smoke",
    ],
}

RETURN_WORDS = [
    "回座", "已回座", "回到座位", "回来了", "到座",
    "back", "im back", "i'm back", "iam back", "returned",
    "return", "done", "finish", "finished", "ok", "okay",
]


# =========================
# 输入框下方按钮
# =========================
def get_main_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton("🟢 上班 On"), KeyboardButton("🔴 下班 Off")],
            [KeyboardButton("🍚 吃饭 Eat"), KeyboardButton("🚽 厕所 Toilet")],
            [KeyboardButton("🚬 抽烟 Smoke"), KeyboardButton("✅ 回座 Back")],
            [KeyboardButton("📊 我的 /me"), KeyboardButton("📅 今日 /today")],
            [KeyboardButton("🕘 出勤 /attendance"), KeyboardButton("📤 导出 /export")],
        ],
        resize_keyboard=True,
        one_time_keyboard=False,
        selective=False,
        input_field_placeholder="请选择按钮或直接输入 / Choose a button or type",
    )


# =========================
# 数据库
# =========================
def get_conn():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            chat_id INTEGER NOT NULL,
            action TEXT NOT NULL,
            timeout_minutes INTEGER NOT NULL,
            PRIMARY KEY (chat_id, action)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS chats (
            chat_id INTEGER PRIMARY KEY,
            title TEXT,
            chat_type TEXT,
            registered_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS members (
            chat_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            username TEXT,
            full_name TEXT NOT NULL,
            is_bot INTEGER NOT NULL DEFAULT 0,
            last_seen_at TEXT NOT NULL,
            PRIMARY KEY (chat_id, user_id)
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            username TEXT,
            full_name TEXT NOT NULL,
            period_key TEXT NOT NULL,
            on_duty_time TEXT,
            off_duty_time TEXT,
            work_seconds INTEGER NOT NULL DEFAULT 0,
            status TEXT NOT NULL DEFAULT 'off'
        )
    """)

    cur.execute("""
        CREATE TABLE IF NOT EXISTS checkins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            chat_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            username TEXT,
            full_name TEXT NOT NULL,
            mention_name TEXT,
            action TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT,
            duration_seconds INTEGER,
            period_key TEXT NOT NULL,
            status TEXT NOT NULL,
            warning_sent INTEGER NOT NULL DEFAULT 0,
            timed_out INTEGER NOT NULL DEFAULT 0
        )
    """)

    conn.commit()
    conn.close()


# =========================
# 时间工具
# =========================
def now_dt() -> datetime:
    return datetime.now(LOCAL_TZ)


def fmt_dt(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def format_seconds(total_seconds: int) -> str:
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    if hours > 0:
        return f"{hours}小时{minutes}分钟{seconds}秒"
    if minutes > 0:
        return f"{minutes}分钟{seconds}秒"
    return f"{seconds}秒"


def current_period_key(dt: datetime | None = None) -> str:
    dt = dt or now_dt()
    if dt.time() < RESET_TIME:
        target = dt.date() - timedelta(days=1)
    else:
        target = dt.date()
    return target.strftime("%Y-%m-%d")


def previous_period_key(dt: datetime | None = None) -> str:
    dt = dt or now_dt()
    return current_period_key(dt - timedelta(seconds=1))


def in_checkin_hours(dt: datetime | None = None) -> bool:
    dt = dt or now_dt()
    t = dt.time()
    return t >= CHECK_START or t <= CHECK_END


def in_formal_work_hours(dt: datetime | None = None) -> bool:
    dt = dt or now_dt()
    t = dt.time()
    return t >= WORK_START or t <= WORK_END


def seconds_between(start_iso: str, end_iso: str) -> int:
    start_dt = datetime.fromisoformat(start_iso)
    end_dt = datetime.fromisoformat(end_iso)
    return max(int((end_dt - start_dt).total_seconds()), 0)


def diff_seconds(start_iso: str, end_dt: datetime) -> int:
    start_dt = datetime.fromisoformat(start_iso)
    return max(int((end_dt - start_dt).total_seconds()), 0)


# =========================
# 文本工具
# =========================
def normalize_text(text: str) -> str:
    text = (text or "").strip().lower()
    text = text.replace("’", "'").replace("`", "'")
    text = re.sub(r"[-_.,!?;:，。！？；：/\\()\[\]{}]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def contains_phrase(text: str, phrase: str) -> bool:
    return normalize_text(phrase) in normalize_text(text)


def extract_action(text: str):
    text = normalize_text(text)

    button_map = {
        "🟢 上班 on": "上班",
        "🔴 下班 off": "下班",
        "🍚 吃饭 eat": "吃饭",
        "🚽 厕所 toilet": "上厕所",
        "🚬 抽烟 smoke": "抽烟",
        "✅ 回座 back": "回座",
        "上班 on": "上班",
        "下班 off": "下班",
        "吃饭 eat": "吃饭",
        "上厕所 toilet": "上厕所",
        "厕所 toilet": "上厕所",
        "抽烟 smoke": "抽烟",
        "回座 back": "回座",
    }
    if text in button_map:
        return button_map[text]

    for action, words in ACTION_KEYWORDS.items():
        for word in words:
            if contains_phrase(text, word):
                return action
    return None


def is_return_text(text: str) -> bool:
    text = normalize_text(text)
    if text in {"✅ 回座 back", "回座 back"}:
        return True
    return any(contains_phrase(text, word) for word in RETURN_WORDS)


def safe_text(text: str) -> str:
    return html.escape(text or "")


def mention_name(user) -> str:
    if user.username:
        return f"@{user.username}"
    return user.full_name or str(user.id)


def mention_html(user_id: int, username: str, full_name: str) -> str:
    if username:
        return f"@{html.escape(username)}"
    return f'<a href="tg://user?id={user_id}">{html.escape(full_name)}</a>'


# =========================
# 群与成员登记
# =========================
def ensure_chat(chat_id: int, title: str, chat_type: str):
    conn = get_conn()
    cur = conn.cursor()
    now = now_dt().isoformat()
    cur.execute("""
        INSERT INTO chats(chat_id, title, chat_type, registered_at, updated_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(chat_id) DO UPDATE SET
            title = excluded.title,
            chat_type = excluded.chat_type,
            updated_at = excluded.updated_at
    """, (chat_id, title, chat_type, now, now))
    conn.commit()
    conn.close()


def ensure_member(chat_id: int, user):
    if not user:
        return
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO members(chat_id, user_id, username, full_name, is_bot, last_seen_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(chat_id, user_id) DO UPDATE SET
            username = excluded.username,
            full_name = excluded.full_name,
            is_bot = excluded.is_bot,
            last_seen_at = excluded.last_seen_at
    """, (
        chat_id,
        user.id,
        user.username or "",
        user.full_name or str(user.id),
        1 if user.is_bot else 0,
        now_dt().isoformat(),
    ))
    conn.commit()
    conn.close()


def get_registered_chat_ids():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT chat_id FROM chats WHERE chat_type IN ('group', 'supergroup', 'private')")
    rows = [r[0] for r in cur.fetchall()]
    conn.close()
    return rows


def get_known_members(chat_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT user_id, username, full_name, is_bot
        FROM members
        WHERE chat_id = ?
        ORDER BY full_name
    """, (chat_id,))
    rows = cur.fetchall()
    conn.close()
    return rows


# =========================
# 设置
# =========================
def ensure_default_settings(chat_id: int):
    conn = get_conn()
    cur = conn.cursor()
    for action, minutes in DEFAULT_TIMEOUTS.items():
        cur.execute("""
            INSERT OR IGNORE INTO settings(chat_id, action, timeout_minutes)
            VALUES (?, ?, ?)
        """, (chat_id, action, minutes))
    conn.commit()
    conn.close()


def get_timeout(chat_id: int, action: str) -> int:
    ensure_default_settings(chat_id)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT timeout_minutes
        FROM settings
        WHERE chat_id = ? AND action = ?
    """, (chat_id, action))
    row = cur.fetchone()
    conn.close()
    return int(row["timeout_minutes"]) if row else DEFAULT_TIMEOUTS[action]


# =========================
# attendance
# =========================
def get_attendance(chat_id: int, user_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, on_duty_time, off_duty_time, work_seconds, status
        FROM attendance
        WHERE chat_id = ? AND user_id = ? AND period_key = ?
        ORDER BY id DESC
        LIMIT 1
    """, (chat_id, user_id, period_key))
    row = cur.fetchone()
    conn.close()
    return row


def is_on_duty(chat_id: int, user_id: int, period_key: str) -> bool:
    row = get_attendance(chat_id, user_id, period_key)
    return bool(row and row["status"] == "on")


def create_or_update_on_duty(chat_id: int, user_id: int, username: str, full_name: str, on_duty_time: str, period_key: str):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT id
        FROM attendance
        WHERE chat_id = ? AND user_id = ? AND period_key = ?
        ORDER BY id DESC
        LIMIT 1
    """, (chat_id, user_id, period_key))
    row = cur.fetchone()

    if row:
        cur.execute("""
            UPDATE attendance
            SET username = ?, full_name = ?, on_duty_time = ?, off_duty_time = NULL, work_seconds = 0, status = 'on'
            WHERE id = ?
        """, (username, full_name, on_duty_time, row["id"]))
    else:
        cur.execute("""
            INSERT INTO attendance(chat_id, user_id, username, full_name, period_key, on_duty_time, work_seconds, status)
            VALUES (?, ?, ?, ?, ?, ?, 0, 'on')
        """, (chat_id, user_id, username, full_name, period_key, on_duty_time))

    conn.commit()
    conn.close()


def set_off_duty(chat_id: int, user_id: int, off_duty_time: str, work_seconds: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE attendance
        SET off_duty_time = ?, work_seconds = ?, status = 'off'
        WHERE chat_id = ? AND user_id = ? AND period_key = ?
    """, (off_duty_time, work_seconds, chat_id, user_id, period_key))
    conn.commit()
    conn.close()


def get_attendance_rows(chat_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT full_name, on_duty_time, off_duty_time, work_seconds, status
        FROM attendance
        WHERE chat_id = ? AND period_key = ?
        ORDER BY full_name
    """, (chat_id, period_key))
    rows = cur.fetchall()
    conn.close()
    return rows


def get_on_duty_user_ids(chat_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT user_id
        FROM attendance
        WHERE chat_id = ? AND period_key = ? AND on_duty_time IS NOT NULL
    """, (chat_id, period_key))
    ids = {r[0] for r in cur.fetchall()}
    conn.close()
    return ids


# =========================
# checkins
# =========================
def get_open_checkin(chat_id: int, user_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, action, start_time, mention_name
        FROM checkins
        WHERE chat_id = ? AND user_id = ? AND period_key = ? AND status = 'open'
        ORDER BY id DESC
        LIMIT 1
    """, (chat_id, user_id, period_key))
    row = cur.fetchone()
    conn.close()
    return row


def create_checkin(chat_id: int, user_id: int, username: str, full_name: str, mention: str, action: str, start_time: str, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO checkins(
            chat_id, user_id, username, full_name, mention_name,
            action, start_time, period_key, status
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'open')
    """, (
        chat_id, user_id, username, full_name, mention,
        action, start_time, period_key
    ))
    checkin_id = cur.lastrowid
    conn.commit()
    conn.close()
    return checkin_id


def close_checkin(checkin_id: int, end_time: str, duration_seconds: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE checkins
        SET end_time = ?, duration_seconds = ?, status = 'closed'
        WHERE id = ?
    """, (end_time, duration_seconds, checkin_id))
    conn.commit()
    conn.close()


def cancel_checkin(checkin_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE checkins SET status = 'cancelled' WHERE id = ?", (checkin_id,))
    conn.commit()
    conn.close()


def mark_warning_sent(checkin_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE checkins SET warning_sent = 1 WHERE id = ?", (checkin_id,))
    conn.commit()
    conn.close()


def mark_timed_out(checkin_id: int):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE checkins SET timed_out = 1 WHERE id = ?", (checkin_id,))
    conn.commit()
    conn.close()


def cancel_open_checkins_of_period(chat_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        UPDATE checkins
        SET status = 'cancelled'
        WHERE chat_id = ? AND period_key = ? AND status = 'open'
    """, (chat_id, period_key))
    conn.commit()
    conn.close()


def get_action_count(chat_id: int, user_id: int, action: str, period_key: str) -> int:
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) AS c
        FROM checkins
        WHERE chat_id = ?
          AND user_id = ?
          AND action = ?
          AND period_key = ?
          AND status IN ('open', 'closed')
    """, (chat_id, user_id, action, period_key))
    row = cur.fetchone()
    conn.close()
    return int(row["c"]) if row else 0


def get_user_break_summary(chat_id: int, user_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()

    cur.execute("""
        SELECT action, COUNT(*) AS cnt, COALESCE(SUM(duration_seconds), 0) AS seconds
        FROM checkins
        WHERE chat_id = ?
          AND user_id = ?
          AND period_key = ?
          AND status = 'closed'
        GROUP BY action
    """, (chat_id, user_id, period_key))
    rows = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*) AS cnt, COALESCE(SUM(duration_seconds), 0) AS seconds
        FROM checkins
        WHERE chat_id = ?
          AND user_id = ?
          AND period_key = ?
          AND status = 'closed'
    """, (chat_id, user_id, period_key))
    total = cur.fetchone()

    conn.close()

    per_action = {a: {"count": 0, "seconds": 0} for a in BREAK_ACTIONS}
    for row in rows:
        per_action[row["action"]] = {
            "count": row["cnt"],
            "seconds": row["seconds"],
        }

    return per_action, int(total["cnt"] or 0), int(total["seconds"] or 0)


def get_stats(chat_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT full_name, action, COUNT(*) AS cnt, COALESCE(SUM(duration_seconds), 0) AS total_seconds
        FROM checkins
        WHERE chat_id = ?
          AND period_key = ?
          AND status = 'closed'
        GROUP BY full_name, action
        ORDER BY full_name, action
    """, (chat_id, period_key))
    rows = cur.fetchall()
    conn.close()
    return rows


def get_records(chat_id: int, period_key: str):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT full_name, action, start_time, end_time, duration_seconds, status, timed_out
        FROM checkins
        WHERE chat_id = ? AND period_key = ?
        ORDER BY id ASC
    """, (chat_id, period_key))
    rows = cur.fetchall()
    conn.close()
    return rows


# =========================
# 展示文本
# =========================
def format_stats(rows, period_key: str):
    if not rows:
        return f"📅 <b>{period_key} 周期离岗统计 / Break Summary</b>\n\n暂无记录 / No records."

    grouped = defaultdict(dict)
    for row in rows:
        grouped[row["full_name"]][row["action"]] = {
            "count": row["cnt"],
            "seconds": row["total_seconds"],
        }

    lines = [f"📅 <b>{period_key} 周期离岗统计 / Break Summary</b>\n"]
    for name, data in grouped.items():
        lines.append(f"👤 <b>{safe_text(name)}</b>")
        for action in BREAK_ACTIONS:
            item = data.get(action, {"count": 0, "seconds": 0})
            lines.append(f"• {action}：{item['count']} 次 / times，合计 {format_seconds(item['seconds'])}")
        lines.append("")
    return "\n".join(lines).strip()


def format_attendance(rows, period_key: str):
    if not rows:
        return f"🕘 <b>{period_key} 周期上下班 / Attendance</b>\n\n暂无记录 / No records."

    lines = [f"🕘 <b>{period_key} 周期上下班 / Attendance</b>\n"]
    for row in rows:
        on_show = row["on_duty_time"][11:19] if row["on_duty_time"] else "-"
        off_show = row["off_duty_time"][11:19] if row["off_duty_time"] else "-"
        status_show = "在岗 / On duty" if row["status"] == "on" else "已下班 / Off duty"
        lines.append(
            f"👤 <b>{safe_text(row['full_name'])}</b>\n"
            f"• 上班 / On：{on_show}\n"
            f"• 下班 / Off：{off_show}\n"
            f"• 上班时长 / Work：{format_seconds(row['work_seconds'] or 0)}\n"
            f"• 状态 / Status：{status_show}\n"
        )
    return "\n".join(lines).strip()


def format_me(full_name: str, per_action: dict, total_count: int, total_seconds: int, attendance, period_key: str):
    on_show = "-"
    off_show = "-"
    work_seconds = 0
    status_show = "未上班 / Not checked in"

    if attendance:
        on_show = attendance["on_duty_time"][11:19] if attendance["on_duty_time"] else "-"
        off_show = attendance["off_duty_time"][11:19] if attendance["off_duty_time"] else "-"
        work_seconds = attendance["work_seconds"] or 0
        status_show = "在岗 / On duty" if attendance["status"] == "on" else "已下班 / Off duty"

    return (
        f"👤 <b>{safe_text(full_name)} - {period_key} 我的离岗明细 / My Break Summary</b>\n\n"
        f"🕘 上班 / On：{on_show}\n"
        f"🕘 下班 / Off：{off_show}\n"
        f"⏱ 上班时长 / Work：{format_seconds(work_seconds)}\n"
        f"📊 离岗总次数 / Total breaks：{total_count} 次\n"
        f"⌛ 离岗总时长 / Total time：{format_seconds(total_seconds)}\n\n"
        f"🍚 吃饭 / Eat：{per_action['吃饭']['count']} 次，{format_seconds(per_action['吃饭']['seconds'])}\n"
        f"🚽 上厕所 / Toilet：{per_action['上厕所']['count']} 次，{format_seconds(per_action['上厕所']['seconds'])}\n"
        f"🚬 抽烟 / Smoke：{per_action['抽烟']['count']} 次，{format_seconds(per_action['抽烟']['seconds'])}\n"
        f"📍 当前状态 / Status：{status_show}"
    )


def build_excel(chat_id: int, period_key: str) -> BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "上下班日报"
    ws1.append(["姓名", "上班时间", "下班时间", "上班时长(秒)", "状态"])
    for c in ws1[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row in get_attendance_rows(chat_id, period_key):
        ws1.append([
            row["full_name"],
            row["on_duty_time"] or "",
            row["off_duty_time"] or "",
            row["work_seconds"] or 0,
            "在岗" if row["status"] == "on" else "已下班/离岗",
        ])

    ws2 = wb.create_sheet("离岗明细")
    ws2.append(["姓名", "项目", "开始时间", "结束时间", "用时(秒)", "状态", "是否超时"])
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row in get_records(chat_id, period_key):
        ws2.append([
            row["full_name"],
            row["action"],
            row["start_time"],
            row["end_time"] or "",
            row["duration_seconds"] if row["duration_seconds"] is not None else "",
            row["status"],
            "是" if row["timed_out"] else "否",
        ])

    ws3 = wb.create_sheet("个人汇总")
    ws3.append(["姓名", "项目", "次数", "总时长(秒)"])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row in get_stats(chat_id, period_key):
        ws3.append([row["full_name"], row["action"], row["cnt"], row["total_seconds"]])

    for ws in [ws1, ws2, ws3]:
        for col in "ABCDEFGH":
            ws.column_dimensions[col].width = 22

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# =========================
# 通用发送
# =========================
async def send_reply(update: Update, text: str):
    await update.effective_chat.send_message(
        text,
        parse_mode=ParseMode.HTML,
        reply_markup=get_main_keyboard(),
    )


# =========================
# 权限
# =========================
async def is_group_admin(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    chat = update.effective_chat
    user = update.effective_user

    if chat.type == ChatType.PRIVATE:
        return True

    member = await context.bot.get_chat_member(chat.id, user.id)
    return member.status in ("creator", "administrator")


async def get_admin_mentions(chat_id: int, context: ContextTypes.DEFAULT_TYPE) -> str:
    admins = await context.bot.get_chat_administrators(chat_id)
    items = []
    for admin in admins:
        u = admin.user
        if u.is_bot:
            continue
        items.append(mention_html(u.id, u.username or "", u.full_name or str(u.id)))
    return " ".join(items) if items else "管理员"


# =========================
# 定时任务
# =========================
def ensure_scheduled_for_chat(app: Application, chat_id: int):
    scheduled = app.bot_data.setdefault("scheduled_chat_ids", set())
    if chat_id in scheduled:
        return

    app.job_queue.run_daily(
        auto_export_job,
        time=time(12, 0, tzinfo=LOCAL_TZ),
        days=(0, 1, 2, 3, 4, 5, 6),
        data={"chat_id": chat_id},
        name=f"auto_export_{chat_id}",
    )

    app.job_queue.run_daily(
        absence_check_job,
        time=time(21, 10, tzinfo=LOCAL_TZ),
        days=(0, 1, 2, 3, 4, 5, 6),
        data={"chat_id": chat_id},
        name=f"absence_check_{chat_id}",
    )

    scheduled.add(chat_id)


async def post_init(app: Application):
    for chat_id in get_registered_chat_ids():
        ensure_scheduled_for_chat(app, chat_id)


async def auto_export_job(context: ContextTypes.DEFAULT_TYPE):
    chat_id = context.job.data["chat_id"]
    period_key = previous_period_key(now_dt())

    cancel_open_checkins_of_period(chat_id, period_key)

    bio = build_excel(chat_id, period_key)
    filename = f"打卡记录_{period_key}.xlsx"

    await context.bot.send_document(
        chat_id=chat_id,
        document=InputFile(bio, filename=filename),
        caption=f"📄 {period_key} 打卡记录 / Report exported",
        reply_markup=get_main_keyboard(),
    )


async def absence_check_job(context: ContextTypes.DEFAULT_TYPE):
    chat_id = context.job.data["chat_id"]
    period_key = current_period_key()

    known_members = get_known_members(chat_id)
    on_duty_ids = get_on_duty_user_ids(chat_id, period_key)

    missing = []
    for row in known_members:
        if row["is_bot"] == 1:
            continue
        if row["user_id"] in on_duty_ids:
            continue
        missing.append(mention_html(row["user_id"], row["username"] or "", row["full_name"]))

    if not missing:
        return

    text = (
        f"🚨 <b>21:10 未打上班卡名单 / Missing Check-in List</b>\n\n"
        f"周期 / Period：<b>{period_key}</b>\n"
        + "\n".join(missing)
    )
    await context.bot.send_message(
        chat_id=chat_id,
        text=text,
        parse_mode=ParseMode.HTML,
        reply_markup=get_main_keyboard(),
    )


async def timeout_warning_job(context: ContextTypes.DEFAULT_TYPE):
    data = context.job.data
    chat_id = data["chat_id"]
    user_id = data["user_id"]
    checkin_id = data["checkin_id"]
    action = data["action"]
    mention = data["mention"]
    period_key = data["period_key"]

    row = get_open_checkin(chat_id, user_id, period_key)
    if not row:
        return
    if row["id"] != checkin_id or row["action"] != action:
        return

    mark_warning_sent(checkin_id)

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            f"⚠️ <b>即将超时提醒 / Almost timeout</b>\n\n"
            f"{safe_text(mention)}\n"
            f"项目 / Action：<b>{action}</b>\n"
            f"请尽快回座 / Please go back and send：<b>回座 / BACK / done</b>"
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=get_main_keyboard(),
    )


async def timeout_over_job(context: ContextTypes.DEFAULT_TYPE):
    data = context.job.data
    chat_id = data["chat_id"]
    user_id = data["user_id"]
    checkin_id = data["checkin_id"]
    action = data["action"]
    mention = data["mention"]
    period_key = data["period_key"]
    timeout_minutes = data["timeout_minutes"]

    row = get_open_checkin(chat_id, user_id, period_key)
    if not row:
        return
    if row["id"] != checkin_id or row["action"] != action:
        return

    mark_timed_out(checkin_id)
    admins = await get_admin_mentions(chat_id, context)

    await context.bot.send_message(
        chat_id=chat_id,
        text=(
            f"🚨 <b>离岗超时 / Break timeout</b>\n\n"
            f"{safe_text(mention)}\n"
            f"项目 / Action：<b>{action}</b>\n"
            f"已超过 / Exceeded：<b>{timeout_minutes}</b> 分钟\n\n"
            f"{admins}"
        ),
        parse_mode=ParseMode.HTML,
        reply_markup=get_main_keyboard(),
    )


# =========================
# 命令
# =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat = update.effective_chat
    user = update.effective_user

    ensure_chat(chat.id, chat.title or chat.full_name or str(chat.id), chat.type)
    ensure_member(chat.id, user)
    ensure_default_settings(chat.id)
    ensure_scheduled_for_chat(context.application, chat.id)

    msg = (
        "<b>泰国时间企业打卡机器人 / Thailand Enterprise Check-in Bot</b>\n\n"
        "功能 / Features：\n"
        "• 可打卡时间 / Check-in allowed：20:00 - 次日12:00（泰国时间）\n"
        "• 正式上下班时间 / Formal work hours：21:00 - 次日10:00（泰国时间）\n"
        "• 每天中午12:00自动导出上一周期表格\n"
        "• 每晚21:10自动提醒未打上班卡成员（已登记成员）\n"
        "• 吃饭15分钟 / 上厕所10分钟 / 抽烟10分钟\n"
        "• 快超时先提醒本人，超时后 @管理员\n"
        "• 导出按钮仅管理员可用\n\n"
        "命令 / Commands：\n"
        "/today 查看本周期离岗统计\n"
        "/attendance 查看本周期上下班\n"
        "/me 查看我的离岗次数和时长\n"
        "/status 查看当前配置\n"
        "/export 导出日报\n"
        "/cancel 取消当前离岗"
    )
    await send_reply(update, msg)


async def today_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    period_key = current_period_key()
    rows = get_stats(update.effective_chat.id, period_key)
    await send_reply(update, format_stats(rows, period_key))


async def attendance_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    period_key = current_period_key()
    rows = get_attendance_rows(update.effective_chat.id, period_key)
    await send_reply(update, format_attendance(rows, period_key))


async def me_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    full_name = update.effective_user.full_name or str(user_id)
    period_key = current_period_key()

    per_action, total_count, total_seconds = get_user_break_summary(chat_id, user_id, period_key)
    att = get_attendance(chat_id, user_id, period_key)

    await send_reply(update, format_me(full_name, per_action, total_count, total_seconds, att, period_key))


async def status_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    ensure_default_settings(chat_id)
    lines = [
        "<b>当前配置 / Current Config</b>\n",
        "时区 / Timezone：Asia/Bangkok",
        f"可打卡时间 / Check-in allowed：{CHECK_START.strftime('%H:%M')} - 次日{CHECK_END.strftime('%H:%M')}",
        f"正式时间 / Formal work：{WORK_START.strftime('%H:%M')} - 次日{WORK_END.strftime('%H:%M')}",
        f"周期切换 / Reset：每天 {RESET_TIME.strftime('%H:%M')}",
        f"未打卡检查 / Missing check：每天 {ABSENCE_CHECK_TIME.strftime('%H:%M')}",
        f"导出仅管理员 / Export admin only：{'开启 / ON' if EXPORT_ADMIN_ONLY else '关闭 / OFF'}\n",
        "离岗时限 / Break limits：",
    ]
    for action in BREAK_ACTIONS:
        lines.append(f"• {action}：{get_timeout(chat_id, action)} 分钟")
    await send_reply(update, "\n".join(lines))


async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if EXPORT_ADMIN_ONLY and not await is_group_admin(update, context):
        await send_reply(update, "只有管理员可以导出日报 / Only admins can export reports.")
        return

    period_key = current_period_key()
    chat_id = update.effective_chat.id
    bio = build_excel(chat_id, period_key)
    filename = f"打卡记录_{period_key}.xlsx"

    await update.effective_chat.send_document(
        document=InputFile(bio, filename=filename),
        caption=f"📄 {period_key} 打卡日报 / Report",
        reply_markup=get_main_keyboard(),
    )


async def cancel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    period_key = current_period_key()

    row = get_open_checkin(chat_id, user_id, period_key)
    if not row:
        await send_reply(update, "你当前没有进行中的离岗 / No active break now.")
        return

    cancel_checkin(row["id"])
    await send_reply(update, f"已取消当前离岗 / Cancelled：{row['action']}")


# =========================
# 核心逻辑
# =========================
async def process_action(update: Update, context: ContextTypes.DEFAULT_TYPE, text: str):
    chat = update.effective_chat
    user = update.effective_user
    chat_id = chat.id
    user_id = user.id
    username = user.username or ""
    full_name = user.full_name or str(user.id)
    mention = mention_name(user)
    period_key = current_period_key()

    ensure_chat(chat.id, chat.title or chat.full_name or str(chat.id), chat.type)
    ensure_member(chat.id, user)
    ensure_default_settings(chat.id)
    ensure_scheduled_for_chat(context.application, chat.id)

    # 回座
    if is_return_text(text):
        open_row = get_open_checkin(chat_id, user_id, period_key)
        if not open_row:
            await send_reply(update, "你当前没有进行中的离岗 / No active break.")
            return

        end_dt = now_dt()
        seconds = diff_seconds(open_row["start_time"], end_dt)
        close_checkin(open_row["id"], end_dt.isoformat(), seconds)

        per_action, total_count, total_seconds = get_user_break_summary(chat_id, user_id, period_key)

        await send_reply(
            update,
            (
                f"✅ <b>回座成功 / BACK TO WORK</b>\n\n"
                f"👤 {safe_text(full_name)}\n"
                f"📌 项目 / Action：{open_row['action']}\n"
                f"🕒 本次离岗 / This break：{format_seconds(seconds)}\n"
                f"📊 本周期离岗总次数 / Total breaks：{total_count} 次\n"
                f"⌛ 本周期离岗总时长 / Total time：{format_seconds(total_seconds)}\n"
                f"📍 回座时间 / Back at：{fmt_dt(end_dt)}"
            ),
        )
        return

    action = extract_action(text)
    if not action:
        return

    # 上下班时间限制：只限制可打卡时间
    if action in ["上班", "下班"] and not in_checkin_hours():
        await send_reply(
            update,
            (
                "❌ <b>当前不在可打卡时间 / Not in allowed check-in time</b>\n\n"
                f"🕒 可打卡时间 / Allowed：{CHECK_START.strftime('%H:%M')} - 次日{CHECK_END.strftime('%H:%M')}（TH）"
            )
        )
        return

    # 上班
    if action == "上班":
        if is_on_duty(chat_id, user_id, period_key):
            await send_reply(update, "你本周期已经是上班状态 / You are already on duty in this period.")
            return

        now = now_dt()
        create_or_update_on_duty(chat_id, user_id, username, full_name, now.isoformat(), period_key)

        warning = ""
        if not in_formal_work_hours():
            warning = "\n⚠️ 非正式上班时间 / Outside formal work hours (21:00 - 10:00)"

        await send_reply(
            update,
            (
                f"🟢 <b>上班成功 / ON DUTY</b>\n\n"
                f"👤 {safe_text(full_name)}\n"
                f"🕒 时间 / Time：{fmt_dt(now)}\n"
                f"📅 周期 / Period：{period_key}\n"
                f"⏰ 正式时间 / Formal work hours：21:00 - 次日10:00"
                f"{warning}"
            ),
        )
        return

    # 下班
    if action == "下班":
        if not is_on_duty(chat_id, user_id, period_key):
            await send_reply(update, "你当前不是上班状态 / You are not on duty now.")
            return

        open_row = get_open_checkin(chat_id, user_id, period_key)
        if open_row:
            await send_reply(update, f"你还有未结束离岗 / Active break：{open_row['action']}\n请先发送：回座 / BACK / done")
            return

        att = get_attendance(chat_id, user_id, period_key)
        now = now_dt()
        work_seconds = 0
        if att and att["on_duty_time"]:
            work_seconds = seconds_between(att["on_duty_time"], now.isoformat())

        set_off_duty(chat_id, user_id, now.isoformat(), work_seconds, period_key)
        per_action, total_count, total_seconds = get_user_break_summary(chat_id, user_id, period_key)

        await send_reply(
            update,
            (
                f"🔴 <b>下班成功 / OFF DUTY</b>\n\n"
                f"👤 {safe_text(full_name)}\n"
                f"🕒 时间 / Time：{fmt_dt(now)}\n"
                f"⏱ 本周期上班时长 / Work time：{format_seconds(work_seconds)}\n"
                f"📊 本周期离岗总次数 / Total breaks：{total_count} 次\n"
                f"🧾 吃饭 / Eat：{per_action['吃饭']['count']} 次，{format_seconds(per_action['吃饭']['seconds'])}\n"
                f"🧾 上厕所 / Toilet：{per_action['上厕所']['count']} 次，{format_seconds(per_action['上厕所']['seconds'])}\n"
                f"🧾 抽烟 / Smoke：{per_action['抽烟']['count']} 次，{format_seconds(per_action['抽烟']['seconds'])}\n"
                f"⌛ 本周期离岗总时长 / Total break time：{format_seconds(total_seconds)}"
            ),
        )
        return

    # 离岗
    if action in BREAK_ACTIONS:
        if not is_on_duty(chat_id, user_id, period_key):
            await send_reply(update, "请先上班打卡 / Please check in first.")
            return

        open_row = get_open_checkin(chat_id, user_id, period_key)
        if open_row:
            await send_reply(update, f"你当前还有未结束离岗 / Active break：{open_row['action']}\n请先发送：回座 / BACK / done")
            return

        previous_count = get_action_count(chat_id, user_id, action, period_key)
        current_count = previous_count + 1
        _, total_break_count, total_break_seconds = get_user_break_summary(chat_id, user_id, period_key)

        start_dt = now_dt()
        timeout_minutes = get_timeout(chat_id, action)

        checkin_id = create_checkin(
            chat_id=chat_id,
            user_id=user_id,
            username=username,
            full_name=full_name,
            mention=mention,
            action=action,
            start_time=start_dt.isoformat(),
            period_key=period_key,
        )

        warning_delay = max(timeout_minutes * 60 - WARNING_BEFORE_SECONDS, 5)

        context.job_queue.run_once(
            timeout_warning_job,
            when=warning_delay,
            data={
                "chat_id": chat_id,
                "user_id": user_id,
                "checkin_id": checkin_id,
                "action": action,
                "mention": mention,
                "period_key": period_key,
            },
            name=f"warn_{chat_id}_{user_id}_{checkin_id}",
        )

        context.job_queue.run_once(
            timeout_over_job,
            when=timeout_minutes * 60,
            data={
                "chat_id": chat_id,
                "user_id": user_id,
                "checkin_id": checkin_id,
                "action": action,
                "mention": mention,
                "period_key": period_key,
                "timeout_minutes": timeout_minutes,
            },
            name=f"over_{chat_id}_{user_id}_{checkin_id}",
        )

        await send_reply(
            update,
            (
                f"🟡 <b>离岗成功 / BREAK STARTED</b>\n\n"
                f"👤 {safe_text(full_name)}\n"
                f"📌 项目 / Action：{action}\n"
                f"🕒 开始时间 / Start：{fmt_dt(start_dt)}\n"
                f"🔢 本周期该项目第 / This action No.：{current_count} 次\n"
                f"📊 本周期离岗总次数 / Total breaks：{total_break_count + 1} 次（含本次进行中）\n"
                f"⌛ 已完成离岗总时长 / Finished break time：{format_seconds(total_break_seconds)}\n"
                f"⏰ 超时上限 / Limit：{timeout_minutes} 分钟\n\n"
                f"👉 返回 / Back：<b>回座 / BACK / done</b>"
            ),
        )
        return


# =========================
# 文本处理
# =========================
async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message:
        return
    chat = update.effective_chat
    user = update.effective_user
    ensure_chat(chat.id, chat.title or chat.full_name or str(chat.id), chat.type)
    ensure_member(chat.id, user)

    text = (update.message.text or "").strip()
    await process_action(update, context, text)


# =========================
# 主程序
# =========================
def main():
    init_db()

    defaults = Defaults(tzinfo=LOCAL_TZ)
    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .defaults(defaults)
        .post_init(post_init)
        .build()
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("today", today_command))
    app.add_handler(CommandHandler("attendance", attendance_command))
    app.add_handler(CommandHandler("me", me_command))
    app.add_handler(CommandHandler("status", status_command))
    app.add_handler(CommandHandler("export", export_command))
    app.add_handler(CommandHandler("cancel", cancel_command))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    logger.info("Thailand enterprise checkin bot is running...")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
